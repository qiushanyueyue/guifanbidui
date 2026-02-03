import openpyxl
import re
from typing import Optional, Dict
import os
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelLoader:
    _instance = None
    _standards_map: Dict[str, dict] = {} # Key: normalized code
    _name_map: Dict[str, dict] = {}      # Key: normalized name

    def __new__(cls, file_path: str = None):
        if cls._instance is None:
            cls._instance = super(ExcelLoader, cls).__new__(cls)
            # Default to the user's provided path if none specified
            # Use local copy if available, fallback to volume
            # When running from backend/, the file is in current dir
            local_path = "standards_data.xlsx" 
            volume_path = "/Volumes/yue/Download/规范目录库（含网址）.xlsx"
            
            # Check current directory
            if os.path.exists(local_path):
                path_to_load = local_path
            # Check backend directory (if running from root)
            elif os.path.exists(os.path.join("backend", local_path)):
                 path_to_load = os.path.join("backend", local_path)
            # Vercel deployment structure: File might be in the root or parallel
            elif os.path.exists(os.path.join(os.getcwd(), local_path)):
                path_to_load = os.path.join(os.getcwd(), local_path)
            else:
                path_to_load = file_path if file_path else volume_path
            
            # Log the resolved path
            try:
                logger.info(f"Resolved Excel path: {path_to_load} (CWD: {os.getcwd()})")
            except:
                pass
            
            cls._instance.load_data(path_to_load)
        return cls._instance

    def load_data(self, file_path: str):
        try:
            logger.info(f"Loading standards from Excel (using openpyxl): {file_path}")
            
            # Load workbook using openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            
            # Map headers to indices
            header_map = {str(h).strip(): i for i, h in enumerate(headers) if h}
            
            # Check for new format columns
            if '名称' in header_map and '网址' in header_map:
                self._load_new_format(ws, header_map)
            elif '规范名称及编号' in header_map:
                self._load_old_format(ws, header_map)
            else:
                logger.error(f"Unknown Excel format: columns not found. Headers: {headers}")
                
            logger.info(f"Loaded {len(self._standards_map)} codes and {len(self._name_map)} names from Excel.")
            
            # Inject manual overrides
            self._inject_manual_data()
            
            wb.close()
            
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
        finally:
            # Inject manual overrides even if Excel fails
            self._inject_manual_data()

    def _load_new_format(self, ws, header_map):
        name_idx = header_map.get('名称')
        url_idx = header_map.get('网址')
        
        # Iterate from row 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row: continue
            
            content = row[name_idx]
            url = row[url_idx] if url_idx < len(row) else None
            
            if not isinstance(content, str):
                continue
            
            # Parse "《Name》Code" or "Prefix,《Name》Code"
            # Extract Name (inside 《》) and Code (after 》)
            name_match = re.search(r"《(.*?)》", content)
            
            if name_match:
                name = name_match.group(1).strip()
                # Code typically follows the closing bracket
                
                # Heuristic: Find pattern matching code in the whole string
                code_match = re.search(r"([A-Z/]{2,}\s*[A-Z]?\s*\d+(?:\.\d+)?-\d{4})", content)
                code = code_match.group(1).strip() if code_match else ""
                
                entry = {
                    "name": name,
                    "code": code,
                    "full_content": content,
                    "source": "excel",
                    "soujianzhu_url": url
                }
                
                if code:
                    norm_code = self._normalize_code(code)
                    self._standards_map[norm_code] = entry
                
                if name:
                    norm_name = self._normalize_name(name)
                    self._name_map[norm_name] = entry

    def _load_old_format(self, ws, header_map):
        col_idx = header_map.get('规范名称及编号')
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row: continue
            
            content = row[col_idx]
            if not isinstance(content, str):
                continue
            
            match = re.search(r"《(.*?)》(.*)", content)
            if match:
                name = match.group(1).strip()
                code = match.group(2).strip()
                
                entry = {
                    "name": name,
                    "code": code,
                    "full_content": content,
                    "source": "excel",
                    "soujianzhu_url": None
                }
                
                normalized_code = self._normalize_code(code)
                self._standards_map[normalized_code] = entry
                norm_name = self._normalize_name(name)
                self._name_map[norm_name] = entry

    def _normalize_code(self, code: str) -> str:
        """Remove spaces and convert to uppercase for consistent lookups"""
        if not code: return ""
        return re.sub(r"\s+", "", str(code)).upper()

    def _normalize_name(self, name: str) -> str:
        """Remove spaces for consistent name lookups"""
        if not name: return ""
        # First clean the name of artifacts like (共二册)
        clean = self._clean_name_text(str(name))
        return re.sub(r"\s+", "", clean).strip()

    def _clean_name_text(self, name: str) -> str:
        """Clean specific suffixes from name"""
        # Remove (共X册)
        name = re.sub(r"[(（]共.*?[册分卷][)）]", "", name)
        return name

    def search_by_code(self, code: str) -> Optional[dict]:
        """Search for a standard by its code"""
        normalized_query = self._normalize_code(code)
        return self._standards_map.get(normalized_query)
        
    def search_by_name(self, name: str) -> Optional[dict]:
        """Search for a standard by its name"""
        normalized_query = self._normalize_name(name)
        return self._name_map.get(normalized_query)

    def search_fuzzy(self, name: str) -> Optional[dict]:
        """
        Fuzzy search for Soujianzhu links.
        Returns match if:
        1. Query name is contained in Standard name
        2. Standard name is contained in Query name
        """
        if not name or len(name) < 2:
            return None
            
        normalized_query = self._normalize_name(name)
        
        # Iterate all unique standards
        # (This is O(N) but N ~3700 is negligible for Python)
        for std in self._standards_map.values():
            std_name = std.get("name", "")
            if not std_name: continue
            
            norm_std_name = self._normalize_name(std_name)
            
            if normalized_query in norm_std_name or norm_std_name in normalized_query:
                return std
                
        return None

    def _inject_manual_data(self):
        """Inject manually provided standards that are missing from Excel"""
        manual_entries = [
            {
                "name": "地铁设计规范",
                "code": "GB 50157-2013",
                "full_content": "地铁设计规范 GB 50157-2013",
                "source": "manual",
                "soujianzhu_url": "https://www.soujianzhu.cn/NormAndRules/NormContent.aspx?id=2246"
            },
            {
                "name": "钢筋焊接及验收规程",
                "code": "JGJ 18-2012",
                "full_content": "钢筋焊接及验收规程 JGJ 18-2012",
                "source": "manual",
                "soujianzhu_url": "http://www.csres.com/detail/223722.html"
            },
            {
                "name": "民用建筑设计统一标准",
                "code": "GB 50352-2019",
                "full_content": "民用建筑设计统一标准 GB 50352-2019",
                "source": "manual",
                "soujianzhu_url": "https://www.soujianzhu.cn/NormAndRules/NormList.aspx?Key=GB%2050352-2019"
            },
            {
                "name": "无障碍设计规范",
                "code": "GB 50763-2012",
                "full_content": "无障碍设计规范 GB 50763-2012",
                "source": "manual",
                "soujianzhu_url": "https://www.soujianzhu.cn/NormAndRules/NormList.aspx?Key=GB%2050763-2012"
            },
            {
                "name": "办公建筑设计标准",
                "code": "JGJ/T 67-2019",
                "full_content": "办公建筑设计标准 JGJ/T 67-2019",
                "source": "manual",
                "soujianzhu_url": "https://www.soujianzhu.cn/NormAndRules/NormList.aspx?Key=JGJ%2FT%2067-2019"
            }
        ]
        
        for entry in manual_entries:
            # Add to code map
            if entry["code"]:
                norm_code = self._normalize_code(entry["code"])
                self._standards_map[norm_code] = entry
            
            # Add to name map
            if entry["name"]:
                norm_name = self._normalize_name(entry["name"])
                self._name_map[norm_name] = entry
                
        logger.info(f"Injected {len(manual_entries)} manual standards.")

# Global instance
excel_loader = ExcelLoader()
